# pylint: disable=too-many-statements
"""
Module providing an RESTful endpoint, representing Observer species counts

Copyright (C) 2015-2017 ERT Inc.
"""
import csv
import io
import codecs

import falcon
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

import api.json as json
from api.auth import auth
from api.resources.source import (
    source
    ,parameters as source_parameters
    ,variables
)
from api.resources.source.data import data
from . import (
     parameters
    ,defaults
    ,streaming
)
from api.resources.source.warehouse import warehouse
import sqlalchemy.exc #DatabaseError

__author__ = "Brandon J. Van Vaerenbergh <brandon.vanvaerenbergh@noaa.gov>, "

str_route = 'selection.'
"""
String representing the URI path for this endpoint

E.g., "/selection.{type}"
"""

str_param_type = 'type'
#String representing Falcon uri parameter for desired response format

class Selection:
    """
    A Falcon resource object, representing a FRAM dataset subselection
    """

    def on_get(self, request, resp, **kwargs):
        """
        Falcon resource method, for handling HTTP request GET method

        Falcon request provides: parameters embedded in URL via a keyword args
        dict, as well as convenience class variables falcon.HTTP_*

        FIXME: remove moduel pylint:disable= & refactor this overlong code block!
        """
        # obtain logged in API user ID (if available)
        api_session_user = auth.get_user_id(request)
        # select data
        with warehouse.get_source_model_session() as dwsupport_model:
            sources = source.SourceUtil.get_list_of_data_sources(
                 request.url
                ,auth.get_user_id(request)
                ,dwsupport_model)
            str_dataset_id = get_requested_dataset_id( sources, request, resp, kwargs)
            list_variables_requested_source = variables.get_list_of_variables( str_dataset_id)
            # convert 'datasets' into a list of variables
            list_requested_datasets = parameters.get_requested_datasets( request)
            list_variables_from_datasets = []
            for str_id in list_requested_datasets:
                if str_dataset_id == str_id:
                    list_variables_from_datasets = list_variables_requested_source
                    break
                if str_dataset_id == 'warehouse': #FIXME: refactor this into a source.warehouse function
                    #obtain the 'warehouse' field aliases for each dataset
                    list_source_variables = variables.get_list_of_variables( str_id)
                    for var in list_source_variables:
                        warehouse_utils = api.resources.source.warehouse.warehouse
                        str_alias = warehouse_utils.prefix_field_name( var, str_id)
                        list_variables_from_datasets.append( str_alias)
                else: #error; not a warehouse request & Dataset does not match requested ID
                    raise falcon.HTTPNotFound(description= "Unrecognized dataset: "
                                           + str_id)
            list_requested_variables = parameters.get_requested_variables( request)

            # add default variables
            if len(list_requested_variables) < 1:
                requested_default_query = parameters.get_list_requested_parameter(
                    defaults.PARAMETER_NAME, request)
                try:
                    default_variables = defaults.get_default_variables(
                         requested_default_query
                        ,str_dataset_id
                        ,dwsupport_model)
                except defaults.UndefinedDefaultQuery as error:
                    msg = ("Value {} is not defined for dataset: '{}'"
                           .format(error, str_dataset_id))
                    raise falcon.HTTPInvalidParam(msg, defaults.PARAMETER_NAME)
                except defaults.AmbiguousDefaultQuery as error:
                    msg = "More than one value was specified: {}".format(error)
                    raise falcon.HTTPInvalidParam(msg, defaults.PARAMETER_NAME)
                except defaults.AmbiguousQueryHierarchy as error:
                    raise falcon.HTTPBadRequest( #TODO: add functional test coverage
                        title="Missing Parameter"
                        ,description=(
                           "Selection defaults not clear for"
                           " data source: '{}'."
                           " Selection must specify one or more 'variables='"
                           " selection parameters (or a 'defaults=' parameter"
                           " value from the following list: {})"
                           ).format(str_dataset_id, error)
                    )
                list_requested_variables.extend(default_variables)

            # add variables derived from 'datasets' param
            list_requested_variables.extend( list_variables_from_datasets)
            list_requested_filters = parameters.get_requested_filters( request)
            # process pivot columns parameter
            try:
                pivot_column_variables = parameters.get_requested_pivot_columns(
                    request
                    ,str_dataset_id
                    ,dwsupport_model['tables'])
            except parameters.PivotVariableError as err:
                raise falcon.HTTPInvalidParam(
                    msg=str(err)
                    ,param_name=parameters.ReservedParameterNames.pivot_columns
                ) from err
            # process 'Empty_cells' parameter
            try:
                empty_cell_dimensions = parameters.get_requested_empty_cells(
                    request
                    ,str_dataset_id
                    ,dwsupport_model['tables']
                    ,dwsupport_model['associations']
                )
            except (parameters.EmptyCellsSourceError
                    ,parameters.EmptyCellsDimensionError) as err:
                raise falcon.HTTPInvalidParam(
                    msg=str(err)
                    ,param_name=parameters.ReservedParameterNames.empty_cells
                ) from err
            # retrieve data
            try:
                result_generator = data.get_data(str_dataset_id
                                            ,list_requested_variables
                                            ,list_requested_filters
                                            ,pivot_column_variables
                                            ,empty_cell_dimensions
                                            ,user_id=api_session_user)
            except sqlalchemy.exc.DatabaseError as err:
                raise falcon.HTTPInternalServerError(
                    title='500'
                    ,description="Please try again"
                ) from err
            except data.NoSourceException as err:
                raise falcon.HTTPNotFound(description=("Source '{}' dataset not found:"
                                                       " {}").format(str_dataset_id,err)) from err
            except parameters.FilterVariableError as err:
                #TODO: the bad HTTP parameter not always 'filters',sometimes a user-defined param (implicit-filter)
                #TODO: perhaps parameters should raise two different Exceptions?
                raise falcon.HTTPInvalidParam(str(err), 'filters') from err
            except data.NotAuthorizedException as error:
                raise falcon.HTTPUnauthorized(
                    title='401'
                    ,description=("Selection from sensitive data source '{}'"
                                  " not authorized").format(str_dataset_id)
                ) from error
            str_format_type = get_requested_format_type( kwargs)
            resp.content_type = get_format_http_content_type( str_format_type)
            result_stream = format_result(result_generator, str_format_type)
            chunked_stream = streaming.biggerchunks_stream(result_stream, 4)#2(13.6),3(13),4(
            if str_format_type == 'xlsx':
                byte_stream = chunked_stream #already bytes
            else:
                encoding = 'utf-8'
                if resp.content_type == 'text/csv':
                    encoding = 'utf-8-sig'
                byte_stream = codecs.iterencode(chunked_stream, encoding)
            resp.stream = byte_stream#content

str_format_dict_function = 'function'
#String dict key, for storing references to formatter functions in Format dicts
str_format_dict_http_content = 'content-type'
#String key, for storing Strings representing HTTP Content-Type in Format dicts

def get_format_http_content_type( str_format_type):
    """
    Utility function, convert URI type id into HTTP Content-Type header value

    >>> get_format_http_content_type( 'json')
    'application/json'
    >>> get_format_http_content_type( 'Json')
    'application/json'
    >>> get_format_http_content_type( 'csv')
    'text/csv'
    >>> get_format_http_content_type('xlsx')
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    """
    str_type_lower_trimmed = str_format_type.strip().lower()
    for str_type_key in dict_format_dicts_by_type.keys():
        if str_type_key == str_type_lower_trimmed:
            dict_format = dict_format_dicts_by_type[str_type_key]
            str_http_content_type = dict_format[str_format_dict_http_content]
            return str_http_content_type
    else:
        raise falcon.HTTPNotFound(description= "Unrecognized format type: "
                                               + str_response_format_type)

def format_result(list_result, str_response_format_type='json'):
    """
    Utility function, format a dataset Selection Falcon request

    Exceptions:
    falcon.error.HTTPNotFound -- raised when referenced format type is
        unrecognized.
    
    Returns:
    String generator -- formatted dataset selection

    >>> def test_generator1():
    ...     raise StopIteration()
    ...     yield False #impossible
    >>> out = format_result(test_generator1())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> ''.join(out)#consume iterator & concat returned strings
    '[]'
    >>> def test_generator2():
    ...     yield ('foo', 'bar', 'data')
    ...     yield (1, 2, 42)
    >>> out = format_result(test_generator2())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> ''.join(out)#consume iterator & concat returned strings
    '[{"bar": 2, "data": 42, "foo": 1}]'
    >>> def test_generator3():
    ...     yield ('a', 'z', 'b')
    ...     yield (1, 26, 2)
    ...     yield (3, 28, 4)
    >>> out = format_result(test_generator3())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> ''.join(out)#consume iterator & concat returned strings
    '[{"a": 1, "b": 2, "z": 26}, {"a": 3, "b": 4, "z": 28}]'
    """
    #TODO: implement additional output formats
    try:
        str_type_lower_trimmed = str_response_format_type.strip().lower()
        # get the dict, which contains our formatting function
        dict_format = dict_format_dicts_by_type[str_type_lower_trimmed]
    except KeyError as e:
        raise falcon.HTTPNotFound(description= "Unrecognized format type: "
                                               + str_response_format_type)
    # retrieve the formatting function, from the dict
    function_format = dict_format[str_format_dict_function]
    return function_format(list_result)

def format_result_csv(result_generator):
    """
    Utility function, CSV format a dataset Selection

    Returns:
    String generator -- formatted dataset selection

    >>> def test_generator1():
    ...     raise StopIteration()
    ...     yield False #impossible
    >>> out = format_result_csv(test_generator1())
    >>> ''.join(out)#consume iterator & concat returned strings
    '\\r\\n'
    >>> def test_generator2():
    ...     yield ('foo', 'bar', 'data')
    ...     yield (1, 2, 42)
    >>> out = format_result_csv(test_generator2())
    >>> ''.join(out)#consume iterator & concat returned strings
    '"bar","data","foo"\\r\\n"2","42","1"\\r\\n'
    >>> def test_generator3():
    ...     yield ('a', 'z', 'b')
    ...     yield (1, 26, 2)
    ...     yield (3, 28, 4)
    >>> out = format_result_csv(test_generator3())
    >>> ''.join(out)#consume iterator & concat returned strings
    '"a","b","z"\\r\\n"1","2","26"\\r\\n"3","4","28"\\r\\n'
    >>> def test_generator4():
    ...     yield ('foo', 'bar', 'data')
    >>> out = format_result_csv(test_generator4())
    >>> ''.join(out)#consume iterator & concat returned strings
    '"bar","data","foo"\\r\\n'
    """
    string_stream_output = io.StringIO()
    bool_sort_keys=True
    # extract the first set (header row) from the list of sets& write to buffer
    try:
        tuple_header_unsorted = next(result_generator)
        list_header = list(tuple_header_unsorted)
        if bool_sort_keys:
            list_header.sort()
    except StopIteration: #No results! No formatting needed
        writer = csv.DictWriter( string_stream_output, fieldnames=[], quoting=csv.QUOTE_ALL)
        writer.writeheader()
        yield string_stream_output.getvalue()
        raise StopIteration()

    writer = csv.DictWriter( string_stream_output, fieldnames=list_header, quoting=csv.QUOTE_ALL)
    writer.writeheader()
    yield string_stream_output.getvalue()

    # convert the list of remaining sets to dicts, and write to buffer
    while True:
        set_row = next(result_generator)
        # build a dict; representing this rows's values, plus the header info
        dict_row = {}
        # add each element of the set to dict, using header value as key
        for index_header, str_header in enumerate(tuple_header_unsorted):
            dict_row[str_header] = set_row[index_header]
        # yield new row string
        string_stream_output = io.StringIO()
        writer = csv.DictWriter(string_stream_output, fieldnames=list_header, quoting=csv.QUOTE_ALL)
        writer.writerow(dict_row)
        yield string_stream_output.getvalue()

def format_result_json(result_generator):
    """
    Utility function, JSON format a dataset Selection

    Returns:
    String generator -- formatted dataset selection

    >>> def test_generator1():
    ...     raise StopIteration()
    ...     yield False #impossible
    >>> out = format_result_json(test_generator1())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> ''.join(out)#consume iterator & concat returned strings
    '[]'
    >>> def test_generator2():
    ...     yield ('foo', 'bar', 'data')
    ...     yield (1, 2, 42)
    >>> out = format_result_json(test_generator2())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> ''.join(out)#consume iterator & concat returned strings
    '[{"bar": 2, "data": 42, "foo": 1}]'
    >>> def test_generator3():
    ...     yield ('a', 'z', 'b')
    ...     yield (1, 26, 2)
    ...     yield (3, 28, 4)
    >>> out = format_result_json(test_generator3())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> ''.join(out)#consume iterator & concat returned strings
    '[{"a": 1, "b": 2, "z": 26}, {"a": 3, "b": 4, "z": 28}]'
    """
    bool_sort_keys=True
    # extract the first set (header row) from the list of sets
    try:
        tuple_header = next(result_generator)
    except StopIteration: #No results! No formatting needed
        string_stream = json.JSONEncoder(sort_keys=bool_sort_keys).iterencode([])
        return string_stream
    # convert the list of remaining sets to a new list of dicts
    dict_generator = json_generator(tuple_header, result_generator)
    #FIXME: empty results look confusing, return 2 lists (header,data?)
    string_stream = json.JSONEncoder(sort_keys=bool_sort_keys,iterable_as_array=True).iterencode(dict_generator)
    return string_stream

def format_result_xlsx(result_generator):
    """
    Utility function, XLSX format a dataset Selection

    Returns:
    String generator -- formatted dataset selection

    >>> from openpyxl import load_workbook
    >>> def test_generator1():
    ...     raise StopIteration()
    ...     yield False #impossible
    >>> out = format_result_xlsx(test_generator1())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> xlsx_stream = io.BytesIO(b''.join(out))#consume iterator
    >>> empty = load_workbook(xlsx_stream, read_only=True)
    >>> all(tab in empty for tab in ['data', 'description'])
    True
    >>> empty['data'].max_column, empty['data'].max_row
    (1, 1)
    >>> def test_generator2():
    ...     yield ('foo', 'bar', 'data')
    ...     yield (1, 2, 42)
    >>> out = format_result_xlsx(test_generator2())
    >>> '__iter__' in dir(out)#check if iterable
    True
    >>> xlsx_stream = io.BytesIO(b''.join(out))#consume iterator
    >>> two_row = load_workbook(xlsx_stream, read_only=True)
    >>> two_row['data'].max_column, two_row['data'].max_row
    (3, 2)
    >>> def test_generator3():
    ...     yield ('a', 'z', 'b')
    ...     yield (1, 26, 2)
    ...     yield (3, 28, 4)
    >>> out = format_result_xlsx(test_generator3())
    >>> xlsx_stream = io.BytesIO(b''.join(out))#consume iterator
    >>> three = load_workbook(xlsx_stream, read_only=True)
    >>> # check data & column order
    >>> [[cell.value for cell in row] for row in three['data'].rows]
    [['a', 'b', 'z'], [1, 2, 26], [3, 4, 28]]
    """
    bool_sort_keys=True

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "data"
    description_sheet = workbook.create_sheet("description")
    # TODO: add description metadata

    # extract the first set (header row) from the list of sets& write to sheet
    try:
        tuple_header_unsorted = next(result_generator)
        list_header = list(tuple_header_unsorted)
        if bool_sort_keys:
            list_header.sort()
        data_sheet.append(list_header) #add header row
    except StopIteration: #No results! No formatting needed
        yield save_virtual_workbook(workbook)
        raise StopIteration()
    # convert the list of remaining sets to dicts, and write to buffer
    try:
        while True:
            set_row = next(result_generator)
            # build a dict; representing rows values by openpyxl column number
            dict_row = {}
            # add each element of the set to dict,
            for index_header, str_header in enumerate(tuple_header_unsorted):
                # find which column (starting from 1) heading was used in header row
                workbook_index = list_header.index(str_header)+1
                dict_row[workbook_index] = set_row[index_header]
            data_sheet.append(dict_row)
    except StopIteration:
        yield save_virtual_workbook(workbook)
        raise StopIteration()

def json_generator(tuple_header, row_generator):
    """
    #TODO: document
    """
    while True:
        set_row = next(row_generator)
        # build a dict; representing this rows's values, plus the header info
        dict_row = {}
        # add each element of the set to dict, using header value as key
        for index_header, str_header in enumerate(tuple_header):
            dict_row[str_header] = set_row[index_header]
        yield dict_row

dict_format_dicts_by_type = {
     'json': { str_format_dict_function: format_result_json
             , str_format_dict_http_content: 'application/json'}
    ,'csv' : { str_format_dict_function: format_result_csv
             , str_format_dict_http_content: 'text/csv'}
    ,'xlsx': { str_format_dict_function: format_result_xlsx
              ,str_format_dict_http_content: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
}
#Dictionary, mapping format IDs to dicts that represent format details

def get_requested_dataset_id(sources, request, resp, params):
    """
    Get 'source_id' URL path param for selection, from a Falcon request

    Returns:
    String -- the requested ID

    Exceptions:
    falcon.HTTPNotFound -- raised when no dataset matches the falcon
      request's source_id value, source_id is not found amongst the
      falcon request's parameters, OR source_id is not available for
      selection (is_selectable==False).

    >>> import falcon.testing #Set up fake Falcon app
    >>> tb = falcon.testing.TestBase()
    >>> tb.setUp()
    >>> tr = falcon.testing.TestResource()
    >>> tb.api.add_route(tb.test_route, tr)
    >>> wsgi_iterable = tb.simulate_request(tb.test_route) #populate req
    >>> s = [{'id':'trawl.HaulCharsSat', 'name':'One Bogus Source'
    ...       ,'is_selectable': True}]
    >>> get_requested_dataset_id( s, tr.req, tr.resp, {'source_id': 'trawl.HaulCharsSat'})
    'trawl.HaulCharsSat'
    >>> s.append({'id':'trawl.HaulCharsUnSat', 'name':'A Second Source'
    ...       ,'is_selectable': False})
    >>> select2params = {'source_id': 'trawl.HaulCharsUnSat'}
    >>> get_requested_dataset_id(s, tr.req, tr.resp, select2params)
    Traceback (most recent call last):
      ...
    falcon.errors.HTTPNotFound
    >>> get_requested_dataset_id( s, tr.req, tr.resp, {'no_id': 'SomethingElse'})
    Traceback (most recent call last):
      ...
    falcon.errors.HTTPNotFound
    >>> get_requested_dataset_id( s, tr.req, tr.resp, {'source_id': 'FooData'})
    Traceback (most recent call last):
      ...
    falcon.errors.HTTPNotFound
    """
    # get identifier, indicating the requested dataset
    requested_id = source_parameters.get_requested_dataset_id(sources, request
                                                              ,resp, params)
    for source in sources:
        if source['id'] == requested_id:
            if not source['is_selectable']:
                msg = "Data source not available for Selection: {}"
                raise falcon.HTTPNotFound(description=msg.format(requested_id))
            return source['id'] #Else, return the selectable ID!
    else:
        raise falcon.HTTPNotFound(description= "No data source by that name: "
                                               + requested_id)

def get_requested_format_type( params):
    """
    Extracts 'str_param_type' URL path param, from a Falcon request

    Returns:
    String -- the requested format type

    Exceptions:
    falcon.HTTPNotFound -- module's 'str_param_type' is not found amongst the
    falcon request's parameters

    >>> get_requested_format_type( {'type': 'json'})
    'json'
    >>> get_requested_format_type( {'no_type': 'SomethingElse'})
    Traceback (most recent call last):
      ...
    falcon.errors.HTTPNotFound
    """
    try:
        str_requested = params[str_param_type]
    except KeyError:
        e = falcon.HTTPNotFound(description="No selection 'format' specified in URL")
        raise e
    return str_requested
